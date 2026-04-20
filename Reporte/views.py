from django.shortcuts import render
from django.http import HttpResponse, JsonResponse
from django.views.decorators.http import require_http_methods
from django.utils.dateparse import parse_date
from EntradaSalidas.models import Entrada
from Producto.models import Producto_gb
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.units import inch
from reportlab.lib import colors
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import io

# Create your views here.
def reporte (request):
    return render(request, "reporte.html")

@require_http_methods(["POST"])
def generar_reporte(request):
    """
    Genera un reporte de inventario en PDF o Excel
    """
    try:
        formato = request.POST.get('formato', 'pdf').lower()
        nombre_archivo = request.POST.get('nombre_archivo', 'reporte_inventario')
        incidencias = request.POST.get('incidencias', '')
        observaciones = request.POST.get('observaciones', '')

        # Validar que se seleccione un formato
        if formato not in ['pdf', 'xlsx']:
            return JsonResponse({'error': 'Formato no válido'}, status=400)

        # Obtener datos de la base de datos
        entradas = Entrada.objects.all()
        productos = Producto_gb.objects.all()

        # Calcular totales
        total_entradas = sum(entrada.cantidad for entrada in entradas)
        total_productos = sum(producto.cantidad for producto in productos)

        if formato == 'pdf':
            return generar_pdf(nombre_archivo, entradas, productos, total_entradas,
                            total_productos, incidencias, observaciones)
        else:
            return generar_excel(nombre_archivo, entradas, productos, total_entradas,
                            total_productos, incidencias, observaciones)

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

def generar_pdf(nombre_archivo, entradas, productos, total_entradas, total_productos, incidencias, observaciones):
    """
    Genera un reporte en formato PDF
    """
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []

    # Estilos
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#0000ff'),
        spaceAfter=30,
        alignment=1
    )

    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=12,
        textColor=colors.HexColor('#00c2b8'),
        spaceAfter=10,
        spaceBefore=15
    )

    # Título
    elements.append(Paragraph("Game'sBar 🎮 - Reporte de Inventario", title_style))
    elements.append(Paragraph(f"Fecha de Generación: {datetime.now().strftime('%d/%m/%Y %H:%M')}", styles['Normal']))
    elements.append(Spacer(1, 0.3*inch))

    # Resumen General
    elements.append(Paragraph("RESUMEN GENERAL", heading_style))
    summary_data = [
        ['Concepto', 'Cantidad'],
        ['Total de Entradas (desde proveedores)', str(total_entradas)],
        ['Total de Productos en Matriz', str(total_productos)],
        ['Número de Proveedores', str(len(set(entrada.proveedor for entrada in entradas)))],
    ]

    summary_table = Table(summary_data, colWidths=[3.5*inch, 1.5*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0000ff')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#eef7ff')]),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 0.3*inch))

    # Detalle de Entradas
    elements.append(Paragraph("DETALLE DE ENTRADAS", heading_style))
    entrada_data = [['Producto', 'Cantidad', 'Unidad', 'Proveedor', 'Fecha']]

    for entrada in entradas[:10]:  # Limitar a 10 para no hacer el PDF muy largo
        entrada_data.append([
            entrada.producto.nombre_producto,
            str(entrada.cantidad),
            entrada.unidad_medida,
            entrada.proveedor.nombre_prov,
            entrada.fecha.strftime('%d/%m/%Y')
        ])

    entrada_table = Table(entrada_data, colWidths=[1.5*inch, 0.8*inch, 0.7*inch, 1.2*inch, 1*inch])
    entrada_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0000ff')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#eef7ff')]),
    ]))
    elements.append(entrada_table)
    elements.append(Spacer(1, 0.3*inch))

    # Incidencias
    if incidencias:
        elements.append(Paragraph("INCIDENCIAS DEL PRODUCTO", heading_style))
        elements.append(Paragraph(incidencias, styles['Normal']))
        elements.append(Spacer(1, 0.2*inch))

    # Observaciones
    if observaciones:
        elements.append(Paragraph("OBSERVACIONES GENERALES", heading_style))
        elements.append(Paragraph(observaciones, styles['Normal']))

    doc.build(elements)
    buffer.seek(0)

    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}.pdf"'
    return response

def generar_excel(nombre_archivo, entradas, productos, total_entradas, total_productos, incidencias, observaciones):
    """
    Genera un reporte en formato Excel
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de Inventario"

    # Estilos
    header_fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Título
    ws['A1'] = "Game'sBar 🎮 - Reporte de Inventario"
    ws['A1'].font = Font(bold=True, size=14, color="0000FF")
    ws.merge_cells('A1:E1')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    ws['A2'] = f"Fecha: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws['A2'].font = Font(size=10)
    ws.merge_cells('A2:E2')

    # Resumen General
    ws['A4'] = "RESUMEN GENERAL"
    ws['A4'].font = Font(bold=True, size=11, color="00C2B8")

    ws['A5'] = "Total de Entradas (desde proveedores)"
    ws['B5'] = total_entradas
    ws['A6'] = "Total de Productos en Matriz"
    ws['B6'] = total_productos
    ws['A7'] = "Número de Proveedores"
    ws['B7'] = len(set(entrada.proveedor for entrada in entradas))

    # Detalle de Entradas
    ws['A9'] = "DETALLE DE ENTRADAS"
    ws['A9'].font = Font(bold=True, size=11, color="00C2B8")

    # Encabezados
    headers = ['Producto', 'Cantidad', 'Unidad', 'Proveedor', 'Fecha']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=10, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Datos de entradas
    row = 11
    for entrada in entradas[:100]:  # Limitar a 100 registros
        ws.cell(row=row, column=1).value = entrada.producto.nombre_producto
        ws.cell(row=row, column=2).value = entrada.cantidad
        ws.cell(row=row, column=3).value = entrada.unidad_medida
        ws.cell(row=row, column=4).value = entrada.proveedor.nombre_prov
        ws.cell(row=row, column=5).value = entrada.fecha.strftime('%d/%m/%Y')

        for col in range(1, 6):
            ws.cell(row=row, column=col).border = border
        row += 1

    # Incidencias
    if incidencias:
        incident_row = row + 2
        ws[f'A{incident_row}'] = "INCIDENCIAS DEL PRODUCTO"
        ws[f'A{incident_row}'].font = Font(bold=True, size=11, color="00C2B8")
        ws[f'A{incident_row + 1}'] = incidencias
        ws[f'A{incident_row + 1}'].alignment = Alignment(wrap_text=True)
        ws.merge_cells(f'A{incident_row + 1}:E{incident_row + 1}')

    # Ancho de columnas
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 15

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}.xlsx"'
    return response
